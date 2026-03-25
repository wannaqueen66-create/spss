#!/usr/bin/env python3
from __future__ import annotations

from pathlib import Path
import re
import math
import openpyxl

ROOT = Path('/root/问卷/附图表')

P_EQ_RE = re.compile(r'(?P<prefix>\bp\s*[=＝]\s*)(?P<val>[0-9]*\.?[0-9]+(?:e[-+]?\d+)?)', flags=re.IGNORECASE)
P_LT_RE = re.compile(r'(?P<prefix>\bp\s*[<＜]\s*)(?P<val>[0-9]*\.?[0-9]+(?:e[-+]?\d+)?)', flags=re.IGNORECASE)


def fmt_p_number(val: float, with_p: bool = False, force_lt: bool | None = None) -> str:
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return ''
    if force_lt is None:
        force_lt = val < 0.001
    if force_lt:
        return 'p<0.001' if with_p else '<0.001'
    s = f'{val:.3f}'
    return f'p={s}' if with_p else s


def replace_inline_p(text: str) -> str:
    def repl_eq(m: re.Match[str]) -> str:
        raw = m.group('val')
        try:
            val = float(raw)
        except Exception:
            return m.group(0)
        return fmt_p_number(val, with_p=True)

    def repl_lt(m: re.Match[str]) -> str:
        raw = m.group('val')
        try:
            val = float(raw)
        except Exception:
            return m.group(0)
        return fmt_p_number(val, with_p=True, force_lt=True if val <= 0.001 else None)

    text = P_EQ_RE.sub(repl_eq, text)
    text = P_LT_RE.sub(repl_lt, text)
    return text


def normalize_header(v) -> str:
    if v is None:
        return ''
    return str(v).strip().lower()


def format_workbook(path: Path) -> tuple[bool, int]:
    wb = openpyxl.load_workbook(path)
    changed = 0

    for ws in wb.worksheets:
        p_cols: set[int] = set()
        fdr_cols: set[int] = set()
        header_row = 1
        for cell in ws[header_row]:
            h = normalize_header(cell.value)
            if h == 'p':
                p_cols.add(cell.column)
            elif h == 'fdr' or h == 'p_fdr' or h == 'p adj' or h == 'p_adj':
                fdr_cols.add(cell.column)

        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str):
                    new_v = replace_inline_p(v)
                    if new_v != v:
                        cell.value = new_v
                        changed += 1
                elif isinstance(v, (int, float)):
                    if cell.column in p_cols and ws.cell(row=header_row, column=cell.column).row == header_row and cell.row > header_row:
                        cell.value = fmt_p_number(float(v), with_p=False)
                        changed += 1
                    elif cell.column in fdr_cols and ws.cell(row=header_row, column=cell.column).row == header_row and cell.row > header_row:
                        # keep FDR numeric meaning but normalize to three decimals as display text for consistency
                        cell.value = f'{float(v):.3f}'
                        changed += 1

    if changed:
        wb.save(path)
        return True, changed
    return False, 0


def main() -> int:
    total_files = 0
    total_changes = 0
    for path in sorted(ROOT.rglob('*.xlsx')):
        changed, n = format_workbook(path)
        if changed:
            total_files += 1
            total_changes += n
            print(f'UPDATED\t{path}\tchanges={n}')
    print(f'SUMMARY\tfiles={total_files}\tchanges={total_changes}')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
