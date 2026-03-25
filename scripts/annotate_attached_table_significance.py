#!/usr/bin/env python3
from __future__ import annotations

from pathlib import Path
import math
import re
import openpyxl
from copy import copy

ROOT = Path('/root/问卷/附图表')

P_EQ_RE = re.compile(r'(?P<prefix>\bp\s*[=＝]\s*)(?P<val>[0-9]*\.?[0-9]+(?:e[-+]?\d+)?)', flags=re.IGNORECASE)
P_LT_RE = re.compile(r'(?P<prefix>\bp\s*[<＜]\s*)(?P<val>[0-9]*\.?[0-9]+(?:e[-+]?\d+)?)', flags=re.IGNORECASE)
PLAIN_P_RE = re.compile(r'^(?P<val><0\.001|[0-9]*\.?[0-9]+)$', flags=re.IGNORECASE)


def stars(p: float) -> str:
    if p < 0.001:
        return '***'
    if p < 0.01:
        return '**'
    if p < 0.05:
        return '*'
    return ''


def parse_p(v) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        if isinstance(v, float) and math.isnan(v):
            return None
        return float(v)
    s = str(v).strip()
    if not s or s in {'—', '-', 'ns', 'N/A'}:
        return None
    if s.startswith('<'):
        try:
            return float(s[1:]) / 2.0
        except Exception:
            return 0.0005 if '0.001' in s else None
    try:
        return float(s)
    except Exception:
        return None


def fmt_p_only(p: float) -> str:
    if p < 0.001:
        return '<0.001***'
    return f'{p:.3f}{stars(p)}'


def fmt_p_inline(p: float) -> str:
    if p < 0.001:
        return 'p<0.001***'
    return f'p={p:.3f}{stars(p)}'


def replace_inline_p(text: str) -> str:
    def repl_eq(m: re.Match[str]) -> str:
        try:
            p = float(m.group('val'))
        except Exception:
            return m.group(0)
        return fmt_p_inline(p)

    def repl_lt(m: re.Match[str]) -> str:
        try:
            raw = float(m.group('val'))
        except Exception:
            return m.group(0)
        p = raw / 2.0 if raw <= 0.001 else raw
        return fmt_p_inline(p)

    text = P_EQ_RE.sub(repl_eq, text)
    text = P_LT_RE.sub(repl_lt, text)
    return text


def clear_bold(cell) -> bool:
    if cell.font and cell.font.bold:
        f = copy(cell.font)
        f.bold = False
        cell.font = f
        return True
    return False


def main() -> int:
    files = 0
    changes = 0
    for path in sorted(ROOT.rglob('*.xlsx')):
        wb = openpyxl.load_workbook(path)
        touched = 0
        for ws in wb.worksheets:
            headers = {str(ws.cell(row=1, column=c).value).strip().lower(): c for c in range(1, ws.max_column + 1) if ws.cell(row=1, column=c).value is not None}
            p_like_cols = {c for h, c in headers.items() if h in {'p', 'holm-adjusted p', 'fdr-corrected p'}}

            for row in ws.iter_rows():
                for cell in row:
                    if cell.row > 1 and clear_bold(cell):
                        touched += 1
                    v = cell.value
                    if cell.row > 1 and cell.column in p_like_cols:
                        p = parse_p(v)
                        if p is not None:
                            new_v = fmt_p_only(p)
                            if str(v) != new_v:
                                cell.value = new_v
                                touched += 1
                    elif isinstance(v, str) and ('p=' in v or 'p<' in v or 'p＜' in v):
                        new_v = replace_inline_p(v)
                        if new_v != v:
                            cell.value = new_v
                            touched += 1
        if touched:
            wb.save(path)
            files += 1
            changes += touched
            print(f'UPDATED\t{path}\tchanges={touched}')
    print(f'SUMMARY\tfiles={files}\tchanges={changes}')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
