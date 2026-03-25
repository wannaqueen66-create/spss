#!/usr/bin/env python3
from __future__ import annotations

from pathlib import Path
import openpyxl
import pandas as pd

ROOT = Path('/root/问卷/附图表')

CONSTRUCT_MAP = {
    'S1': ('空间任务支持性评价', 'Appraisal of Task Supportiveness in Space'),
    'S2': ('空间任务支持性评价', 'Appraisal of Task Supportiveness in Space'),
    'S3': ('空间任务支持性评价', 'Appraisal of Task Supportiveness in Space'),
    'S4': ('情感—行为结果评价', 'Affective and Behavioral Outcome Appraisal'),
    'S5': ('情感—行为结果评价', 'Affective and Behavioral Outcome Appraisal'),
    'B1': ('功能性器材要素的认知评价', 'Cognitive Appraisal of Functional Equipment Elements'),
    'B2': ('功能性器材要素的认知评价', 'Cognitive Appraisal of Functional Equipment Elements'),
    'B3': ('功能性器材要素的认知评价', 'Cognitive Appraisal of Functional Equipment Elements'),
}

TARGET_FILES = [
    '表1_所有问卷指标的总体描述性统计.xlsx',
    '表2_S1-S5的LMM_fixed_effects结果.xlsx',
    '表4_B1-B3的描述性统计与模型拟合状态.xlsx',
    '表4_B1-B3的描述性统计与修正后LMM结果.xlsx',
    '表5_S1-S5在不同WWR条件下的描述性统计与pairwise comparisons.xlsx',
    '表5补充_S1-S5的WWR_pairwise完整版.xlsx',
    '表6_S1-S5在不同Complexity条件下的描述性统计与simple effects结果.xlsx',
]


def get_measure_value(row_dict: dict) -> str | None:
    for key in ('Measure', 'Outcome'):
        val = row_dict.get(key)
        if val is not None and str(val).strip() in CONSTRUCT_MAP:
            return str(val).strip()
    return None


def insert_construct_cols_xlsx(path: Path) -> int:
    wb = openpyxl.load_workbook(path)
    changed = 0

    for ws in wb.worksheets:
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        header_to_col = {str(v).strip(): i + 1 for i, v in enumerate(headers) if v is not None}

        measure_col = None
        for key in ('Measure', 'Outcome'):
            if key in header_to_col:
                measure_col = header_to_col[key]
                break
        if measure_col is None:
            continue

        if 'ConstructZH' not in header_to_col:
            ws.insert_cols(1, amount=2)
            ws.cell(row=1, column=1).value = 'ConstructZH'
            ws.cell(row=1, column=2).value = 'ConstructEN'
            changed += 1
        else:
            # refresh positions after possible prior run
            pass

        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        header_to_col = {str(v).strip(): i + 1 for i, v in enumerate(headers) if v is not None}
        measure_col = header_to_col.get('Measure') or header_to_col.get('Outcome')
        zh_col = header_to_col['ConstructZH']
        en_col = header_to_col['ConstructEN']

        for r in range(2, ws.max_row + 1):
            mv = ws.cell(row=r, column=measure_col).value
            if mv is None:
                continue
            mv = str(mv).strip()
            if mv not in CONSTRUCT_MAP:
                continue
            zh, en = CONSTRUCT_MAP[mv]
            if ws.cell(row=r, column=zh_col).value != zh:
                ws.cell(row=r, column=zh_col).value = zh
                changed += 1
            if ws.cell(row=r, column=en_col).value != en:
                ws.cell(row=r, column=en_col).value = en
                changed += 1

    if changed:
        wb.save(path)
    return changed


def update_csv_if_exists(path: Path) -> int:
    if not path.exists():
        return 0
    df = pd.read_csv(path)
    key = None
    for candidate in ('Measure', 'Outcome'):
        if candidate in df.columns:
            key = candidate
            break
    if key is None:
        return 0
    mapped = df[key].astype(str).map(lambda x: CONSTRUCT_MAP.get(x, (None, None))[0])
    if mapped.notna().sum() == 0:
        return 0
    df.insert(0, 'ConstructZH', df[key].astype(str).map(lambda x: CONSTRUCT_MAP.get(x, ('', ''))[0])) if 'ConstructZH' not in df.columns else df.__setitem__('ConstructZH', df[key].astype(str).map(lambda x: CONSTRUCT_MAP.get(x, ('', ''))[0]))
    insert_pos = 1 if 'ConstructZH' in df.columns else 0
    if 'ConstructEN' not in df.columns:
        df.insert(1, 'ConstructEN', df[key].astype(str).map(lambda x: CONSTRUCT_MAP.get(x, ('', ''))[1]))
    else:
        df['ConstructEN'] = df[key].astype(str).map(lambda x: CONSTRUCT_MAP.get(x, ('', ''))[1])
    df.to_csv(path, index=False, encoding='utf-8-sig')
    return 1


def main() -> int:
    total = 0
    for name in TARGET_FILES:
        path = ROOT / name
        if path.exists():
            n = insert_construct_cols_xlsx(path)
            if n:
                total += 1
                print(f'UPDATED\t{path}\tchanges={n}')

    for csv_name in [
        '构念分组图/construct_group_wwr_pairwise.csv',
        '构念分组图/construct_group_complexity_table.csv',
    ]:
        path = ROOT / csv_name
        n = update_csv_if_exists(path)
        if n:
            print(f'UPDATED\t{path}\tchanges=csv')

    print(f'SUMMARY\tfiles={total}')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
